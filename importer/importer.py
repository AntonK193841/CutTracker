from pathlib import Path

from openpyxl import load_workbook

from importer.api_client import CutTrackerApiClient
from importer.models import SpecificationPart


class SpecificationImporter:
    REQUIRED_COLUMNS = {
        "drawing_number",
        "name",
        "width",
        "height",
        "quantity",
        "material_id",
    }

    def __init__(
        self,
        api_client: CutTrackerApiClient,
    ):
        self.api_client = api_client

    def read(
        self,
        file_path: str | Path,
    ) -> list[SpecificationPart]:
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )

        worksheet = workbook.active

        rows = worksheet.iter_rows(
            values_only=True,
        )

        headers = next(rows)

        normalized_headers = [
            str(header).strip().lower()
            for header in headers
        ]

        self._validate_headers(normalized_headers)

        column_indexes = {
            column: normalized_headers.index(column)
            for column in self.REQUIRED_COLUMNS
        }

        parts = []

        for row_number, row in enumerate(rows, start=2):
            if self._is_empty_row(row):
                continue

            try:
                part = SpecificationPart(
                    drawing_number=str(
                        row[column_indexes["drawing_number"]]
                    ).strip(),

                    name=str(
                        row[column_indexes["name"]]
                    ).strip(),

                    width=float(
                        row[column_indexes["width"]]
                    ),

                    height=float(
                        row[column_indexes["height"]]
                    ),

                    quantity=int(
                        row[column_indexes["quantity"]]
                    ),

                    material_id=int(
                        row[column_indexes["material_id"]]
                    ),
                )

                self._validate_part(
                    part,
                    row_number,
                )

                parts.append(part)

            except (TypeError, ValueError) as error:
                raise ValueError(
                    f"Invalid data in row {row_number}: {error}"
                ) from error

        return parts

    def import_file(
        self,
        file_path: str | Path,
    ) -> list[dict]:
        parts = self.read(file_path)

        imported_parts = []

        for part in parts:
            result = self.api_client.create_part(part)
            imported_parts.append(result)

        return imported_parts

    def _validate_headers(
        self,
        headers: list[str],
    ) -> None:
        missing_columns = (
            self.REQUIRED_COLUMNS - set(headers)
        )

        if missing_columns:
            raise ValueError(
                "Missing required columns: "
                + ", ".join(sorted(missing_columns))
            )

    def _validate_part(
        self,
        part: SpecificationPart,
        row_number: int,
    ) -> None:
        if not part.drawing_number:
            raise ValueError(
                f"Row {row_number}: drawing_number is empty"
            )

        if not part.name:
            raise ValueError(
                f"Row {row_number}: name is empty"
            )

        if part.width <= 0:
            raise ValueError(
                f"Row {row_number}: width must be greater than zero"
            )

        if part.height <= 0:
            raise ValueError(
                f"Row {row_number}: height must be greater than zero"
            )

        if part.quantity <= 0:
            raise ValueError(
                f"Row {row_number}: quantity must be greater than zero"
            )

        if part.material_id <= 0:
            raise ValueError(
                f"Row {row_number}: material_id must be greater than zero"
            )

    def _is_empty_row(
        self,
        row: tuple,
    ) -> bool:
        return all(
            value is None
            for value in row
        )